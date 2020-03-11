import openpyxl
import quantities as pq
import pickle
def readText(fileName):
    return " ".join(open(fileName,'r').readlines())
def getDictionaryFromData(fileName): 
    inputBook=openpyxl.load_workbook(fileName,data_only=True)
    sheets=inputBook.sheetnames
    data={}
    for name in sheets:
        sheet=inputBook[name]
        sheetData=[]
        for row in range(3,sheet.max_row+1):
            dictionary={}
            for column in range(1,sheet.max_column+1):
                value=sheet.cell(row=row,column=column).value
                if value is str:
                    value=value.strip()
                if sheet.cell(row=1,column=column).value is not None:
                    unit =sheet.cell(row=1,column=column).value
                    try:
                        value=pq.Quantity(value,str(unit).lower())
                    except LookupError:
                        value=pq.Quantity(value,str(unit))
                    
                dictionary[sheet.cell(row=2,column=column).value.strip()]=value
          
            if dictionary!={}:    
                sheetData.append(dictionary)
        data[name]=sheetData
    return data
def makeWorkSheet(data={}):
    pass
def pickleData(excel,labData):
    data=getDictionaryFromData(excel)
    file=open(labData,"wb")
    pickle.dump(data,file)
    file.close()
def loadData(dataLocation):
    return pickle.load(open(dataLocation,"rb"))
    
